"""
===========================================================
Excel Processing Demo
Libraries:
1. openpyxl
2. pandas
3. pyxlsb
===========================================================
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

FILE = "Covid Dashboard.xlsx"

# ==========================================================
# 1. OPENPYXL
# ==========================================================

print("=" * 70)
print("1. OPENPYXL")
print("=" * 70)

# Load workbook
wb = load_workbook(FILE)

# ----------------------------------------------------------
# Workbook Information
# ----------------------------------------------------------

print("\nWorkbook Information")
print("--------------------")

print("Sheet Names :", wb.sheetnames)

print("Number of Sheets :", len(wb.sheetnames))

print("Active Sheet :", wb.active.title)

# ----------------------------------------------------------
# Read Every Sheet
# ----------------------------------------------------------

for sheet_name in wb.sheetnames:

    ws = wb[sheet_name]

    print("\n" + "="*60)
    print("Sheet :", sheet_name)
    print("="*60)

    print("Rows :", ws.max_row)
    print("Columns :", ws.max_column)

    print("\nFirst 10 Rows\n")

    for row in ws.iter_rows(min_row=1,
                            max_row=min(10, ws.max_row),
                            values_only=True):
        print(row)


# ----------------------------------------------------------
# Read Entire Row
# ----------------------------------------------------------

print("\nRow 2")

for cell in ws[2]:
    print(cell.value)

# ----------------------------------------------------------
# Read Entire Column
# ----------------------------------------------------------

print("\nColumn A")

for cell in ws["A"][:10]:
    print(cell.value)


# Images
# ----------------------------------------------------------

print("\nImages")

if hasattr(ws, "_images"):

    print("Number of Images :", len(ws._images))

    if len(ws._images) == 0:
        print("No Images Found")

else:

    print("No Images Found")

# ----------------------------------------------------------
# Charts
# ----------------------------------------------------------

print("\nCharts")

if hasattr(ws, "_charts"):

    print("Number of Charts :", len(ws._charts))

    for i, chart in enumerate(ws._charts, start=1):

        print(f"Chart {i}")

        print(type(chart).__name__)

else:

    print("No Charts Found")

#
# ==========================================================
# 2. PANDAS
# ==========================================================

print("\n")
print("=" * 70)
print("2. PANDAS")
print("=" * 70)

excel = pd.ExcelFile(FILE)

print("\nSheet Names")

print(excel.sheet_names)

for sheet in excel.sheet_names:

    print("\n"+"="*60)

    print(sheet)

    print("="*60)

    df = pd.read_excel(FILE,
                       sheet_name=sheet)

    print("\nShape")

    print(df.shape)

    print("\nColumns")

    print(df.columns.tolist())

    print("\nHead")

    print(df.head())

    print("\nTail")

    print(df.tail())

    print("\nInfo")

    print(df.info())

    print("\nDescribe")

    print(df.describe(include="all"))

print("\nDone.")